import streamlit as st
import pandas as pd
import sys
from datetime import datetime
import gspread
from google.auth import default

# Authentication check - shared session state from Home page
if 'authenticated' not in st.session_state or not st.session_state.authenticated:
    st.error("🔐 Please log in through the Home page")
    st.stop()

# Add functions to path
sys.path.append('./functions')

import quickbooks
import bigtime
import sheets

st.set_page_config(page_title="Commission Calculator", page_icon="💰", layout="wide")

st.title("💰 Commission Calculator")
st.markdown("Calculate sales commissions from QuickBooks and BigTime data")

# Year selector
col1, col2 = st.columns([1, 3])
with col1:
    year = st.selectbox("Select Year", [2023, 2024, 2025, 2026], index=2)

# Config from secrets
try:
    CONFIG_SHEET_ID = st.secrets["SHEET_CONFIG_ID"]
    REPORTS_FOLDER_ID = st.secrets["REPORTS_FOLDER_ID"]
except:
    # Fallback for Colab (shouldn't happen in Streamlit)
    import credentials
    CONFIG_SHEET_ID = credentials.get("SHEET_CONFIG_ID")
    REPORTS_FOLDER_ID = credentials.get("REPORTS_FOLDER_ID")

if st.button("🚀 Calculate Commissions", type="primary"):
    
    # Collect debug messages
    debug_log = []
    
    # ============================================================
    # PHASE 1: LOAD CONFIGURATION
    # ============================================================
    
    with st.spinner("📋 Loading configuration from Voyage_Global_Config..."):
        rules_df = sheets.read_config(CONFIG_SHEET_ID, "Rules")
        offsets_df = sheets.read_config(CONFIG_SHEET_ID, "Offsets")
        mapping_df = sheets.read_config(CONFIG_SHEET_ID, "Mapping")
        
        if rules_df is None or offsets_df is None or mapping_df is None:
            st.error("❌ Error: Could not load config sheets")
            st.stop()
        
        rules_df = rules_df.rename(columns={'Client': 'Client_or_Resource'})
        
        client_name_map = dict(zip(
            mapping_df[mapping_df['Source_System'] == 'QuickBooks']['Before_Name'],
            mapping_df[mapping_df['Source_System'] == 'QuickBooks']['After_Name']
        ))
        
        debug_log.append(f"✅ Loaded {len(rules_df)} rules, {len(offsets_df)} offsets, {len(client_name_map)} mappings")
        if sheets.should_use_snowflake():
            debug_log.append("❄️ Config: Snowflake")
        else:
            debug_log.append("📊 Config: Google Sheets")
    
    # ============================================================
    # PHASE 2: PULL API DATA
    # ============================================================
    
    with st.spinner("📡 Pulling data from QuickBooks and BigTime..."):
        df_qb_raw = quickbooks.get_consulting_income(year)
        df_bt_raw = bigtime.get_time_report(year)
        
        # Collect QB debug info
        if df_qb_raw is None:
            debug_log.append("❌ QuickBooks API returned None")
        elif df_qb_raw.empty:
            debug_log.append("❌ QuickBooks API returned empty DataFrame")
            debug_log.append("💡 QuickBooks tokens expire every ~100 days. You may need to re-authenticate.")
            debug_log.append("⚠️ This could mean: No transactions found, wrong date range, or connection issue")
        else:
            debug_log.append(f"✅ QB: {len(df_qb_raw)} transactions")
        
        # Collect BT debug info
        if df_bt_raw is None:
            debug_log.append("❌ BigTime API returned None")
        elif df_bt_raw.empty:
            debug_log.append("❌ BigTime API returned empty DataFrame")
        else:
            debug_log.append(f"✅ BT: {len(df_bt_raw)} entries")
        
        if df_qb_raw.empty or df_bt_raw.empty:
            st.error("❌ Error: No data returned from APIs")
            st.info("💡 Check your date range and API credentials")
            st.stop()
        
        debug_log.append(f"✅ Total - QB: ${df_qb_raw['TotalAmount'].astype(float).sum():,.2f} | BT: {len(df_bt_raw)} entries")
    
    # ============================================================
    # PHASE 3: PROCESS QUICKBOOKS DATA (CLIENT COMMISSIONS)
    # ============================================================
    
    with st.spinner("💰 Calculating client commissions..."):
        qb = df_qb_raw.copy()
        qb["TransactionDate"] = pd.to_datetime(qb["TransactionDate"])
        qb["Amount"] = pd.to_numeric(qb["TotalAmount"])
        qb["Year"] = qb["TransactionDate"].dt.year
        
        # Parse client name - split on ":" and take first part
        qb["Client_Raw"] = qb["Customer"].astype(str).str.split(":", n=1).str[0].str.strip()
        qb["Client_Normalized"] = qb["Client_Raw"].replace(client_name_map)
        
        qb_year = qb[qb["Year"] == year].copy()
        
        # Prepare rules
        rules_client = rules_df[rules_df['Rule_Scope'] == 'client'].copy()
        rules_client['Start_Date'] = pd.to_datetime(rules_client['Start_Date'])
        rules_client['End_Date'] = pd.to_datetime(rules_client['End_Date'], errors='coerce')
        
        # Calculate commissions
        commission_records = []
        
        for idx, invoice in qb_year.iterrows():
            client = invoice['Client_Normalized']
            inv_date = invoice['TransactionDate']
            amount = invoice['Amount']
            
            applicable_rules = rules_client[
                (rules_client['Client_or_Resource'] == client) &
                (rules_client['Start_Date'] <= inv_date) &
                ((rules_client['End_Date'].isna()) | (rules_client['End_Date'] >= inv_date))
            ]
            
            for _, rule in applicable_rules.iterrows():
                commission_records.append({
                    'Salesperson': rule['Salesperson'],
                    'Client': client,
                    'Category': rule['Category'],
                    'Invoice_Date': inv_date,
                    'Invoice_Amount': amount,
                    'Commission_Rate': rule['Rate'],
                    'Commission_Amount': amount * rule['Rate'],
                    'Source': 'QuickBooks - Client Commission'
                })
        
        client_commissions = pd.DataFrame(commission_records)
        
        if not client_commissions.empty:
            debug_log.append(f"✅ {len(client_commissions)} client commission entries: ${client_commissions['Commission_Amount'].sum():,.2f}")
        else:
            debug_log.append("⚠️ No client commissions calculated")
    
    # ============================================================
    # PHASE 4: PROCESS BIGTIME DATA (DELIVERY & REFERRAL)
    # ============================================================
    
    with st.spinner("🔨 Calculating delivery & referral commissions..."):
        bt = df_bt_raw.copy()
        
        rules_resource = rules_df[rules_df['Rule_Scope'] == 'resource'].copy()
        rules_resource['Start_Date'] = pd.to_datetime(rules_resource['Start_Date'])
        rules_resource['End_Date'] = pd.to_datetime(rules_resource['End_Date'], errors='coerce')
        
        resource_records = []
        
        # Find columns
        revenue_col = None
        for col_name in ['Billable ($)', 'Revenue_Amount', 'tmchgbillbase']:
            if col_name in bt.columns:
                revenue_col = col_name
                break
        
        date_col = None
        for col_name in ['Date', 'tmdt']:
            if col_name in bt.columns:
                date_col = col_name
                break
        
        staff_col = None
        for col_name in ['Staff Member', 'Staff_Member', 'tmstaffnm']:
            if col_name in bt.columns:
                staff_col = col_name
                break
        
        if revenue_col and date_col and staff_col:
            bt['Revenue'] = pd.to_numeric(bt[revenue_col], errors='coerce')
            bt['Date'] = pd.to_datetime(bt[date_col], errors='coerce')
            bt['Year_Month'] = bt['Date'].dt.to_period('M')
            
            # Get client column if it exists (for delivery commissions)
            client_col = None
            for col_name in ['Client', 'tmclientnm']:
                if col_name in bt.columns:
                    client_col = col_name
                    break
            
            # Apply commission rules to each entry first
            bt_with_rules = []
            for idx, time_entry in bt.iterrows():
                staff = time_entry.get(staff_col, None)
                date = time_entry.get('Date', None)
                revenue = time_entry.get('Revenue', 0)
                year_month = time_entry.get('Year_Month', None)
                client = time_entry.get(client_col, '') if client_col else ''
                
                if not staff or pd.isna(date) or revenue == 0:
                    continue
                
                applicable_rules = rules_resource[
                    (rules_resource['Client_or_Resource'] == staff) &
                    (rules_resource['Start_Date'] <= date) &
                    ((rules_resource['End_Date'].isna()) | (rules_resource['End_Date'] >= date))
                ]
                
                for _, rule in applicable_rules.iterrows():
                    bt_with_rules.append({
                        'Salesperson': rule['Salesperson'],
                        'Resource': staff,
                        'Client': client,
                        'Category': rule['Category'],
                        'Year_Month': year_month,
                        'Revenue': revenue,
                        'Rate': rule['Rate'],
                        'Commission': revenue * rule['Rate']
                    })
            
            if bt_with_rules:
                bt_rules_df = pd.DataFrame(bt_with_rules)
                
                # Aggregate by month
                referral_df = bt_rules_df[bt_rules_df['Category'] == 'Referral Commission']
                delivery_df = bt_rules_df[bt_rules_df['Category'] == 'Delivery Commission']
                
                # Aggregate Referral (by month, per resource)
                if not referral_df.empty:
                    referral_monthly = referral_df.groupby(
                        ['Salesperson', 'Resource', 'Category', 'Year_Month'], 
                        as_index=False
                    ).agg({
                        'Revenue': 'sum',
                        'Commission': 'sum',
                        'Rate': 'first'
                    })
                    
                    referral_monthly['Invoice_Date'] = referral_monthly['Year_Month'].dt.to_timestamp('M')
                    
                    for _, row in referral_monthly.iterrows():
                        resource_records.append({
                            'Salesperson': row['Salesperson'],
                            'Client': row['Resource'],
                            'Category': row['Category'],
                            'Invoice_Date': row['Invoice_Date'],
                            'Invoice_Amount': row['Revenue'],
                            'Commission_Rate': row['Rate'],
                            'Commission_Amount': row['Commission'],
                            'Source': f'BigTime - {row["Category"]} (Monthly)'
                        })
                
                # Aggregate Delivery (by month, per resource, per client)
                if not delivery_df.empty:
                    delivery_monthly = delivery_df.groupby(
                        ['Salesperson', 'Resource', 'Client', 'Category', 'Year_Month'], 
                        as_index=False
                    ).agg({
                        'Revenue': 'sum',
                        'Commission': 'sum',
                        'Rate': 'first'
                    })
                    
                    delivery_monthly['Invoice_Date'] = delivery_monthly['Year_Month'].dt.to_timestamp('M')
                    
                    for _, row in delivery_monthly.iterrows():
                        client_display = f"{row['Resource']} @ {row['Client']}" if row['Client'] else row['Resource']
                        
                        resource_records.append({
                            'Salesperson': row['Salesperson'],
                            'Client': client_display,
                            'Category': row['Category'],
                            'Invoice_Date': row['Invoice_Date'],
                            'Invoice_Amount': row['Revenue'],
                            'Commission_Rate': row['Rate'],
                            'Commission_Amount': row['Commission'],
                            'Source': f'BigTime - {row["Category"]} (Monthly)'
                        })
        
        resource_commissions = pd.DataFrame(resource_records)
        
        if not resource_commissions.empty:
            debug_log.append(f"✅ {len(resource_commissions)} resource commission entries (monthly aggregated): ${resource_commissions['Commission_Amount'].sum():,.2f}")
        else:
            debug_log.append("⚠️ No resource commissions calculated")
    
    # ============================================================
    # PHASE 5: COMBINE & APPLY OFFSETS
    # ============================================================
    
    with st.spinner("🔄 Applying offsets..."):
        all_commissions = pd.concat([client_commissions, resource_commissions], ignore_index=True)
        
        # Process offsets
        offsets_df['Effective_Date'] = pd.to_datetime(offsets_df['Effective_Date'], format='mixed')
        
        def parse_accounting_amount(val):
            if pd.isna(val):
                return 0
            val_str = str(val).strip()
            is_negative = val_str.startswith('(') and val_str.endswith(')')
            val_clean = val_str.replace('(', '').replace(')', '').replace(',', '').strip()
            try:
                amount = float(val_clean)
                return -amount if is_negative else amount
            except:
                return 0
        
        offsets_df['Amount'] = offsets_df['Amount'].apply(parse_accounting_amount)
        offsets_year = offsets_df[offsets_df['Effective_Date'].dt.year == year]
        
        if not offsets_year.empty:
            for _, offset in offsets_year.iterrows():
                offset_record = pd.DataFrame([{
                    'Salesperson': offset['Salesperson'],
                    'Client': 'Offset',
                    'Category': offset['Category'],
                    'Invoice_Date': offset['Effective_Date'],
                    'Invoice_Amount': 0,
                    'Commission_Rate': 0,
                    'Commission_Amount': offset['Amount'],
                    'Source': f"Offset - {offset.get('Note', '')}"
                }])
                all_commissions = pd.concat([all_commissions, offset_record], ignore_index=True)
            
            debug_log.append(f"✅ Applied {len(offsets_year)} offsets: ${offsets_year['Amount'].sum():,.2f}")
    
    # ============================================================
    # PHASE 6: CALCULATE SUMMARIES
    # ============================================================
    
    final_summary = all_commissions.groupby('Salesperson', as_index=False).agg({
        'Commission_Amount': 'sum'
    }).round(2)
    final_summary.columns = ['Salesperson', 'Total_Commission']
    
    # Calculate Total Due (only positive amounts)
    final_summary['Total_Due'] = final_summary['Total_Commission'].apply(lambda x: max(0, x))
    
    category_summary = all_commissions.groupby(['Salesperson', 'Category'], as_index=False).agg({
        'Commission_Amount': 'sum'
    }).round(2)
    
    revenue_by_client = qb_year.groupby('Client_Normalized').agg({
        'Amount': 'sum',
        'TransactionDate': 'count'
    }).rename(columns={'Amount': 'Total_Revenue', 'TransactionDate': 'Transactions'})
    revenue_by_client = revenue_by_client.sort_values('Total_Revenue', ascending=False)
    
    # Store in session state for email
    st.session_state.commission_report_data = {
        'year': year,
        'final_summary': final_summary,
        'category_summary': category_summary,
        'all_commissions': all_commissions,
        'revenue_by_client': revenue_by_client
    }
    
    debug_log.append("✅ Calculations complete!")
    
    # ============================================================
    # PHASE 7: DISPLAY DEBUG LOG IN EXPANDER
    # ============================================================
    
    with st.expander("🔍 Debug Log", expanded=False):
        for msg in debug_log:
            if msg.startswith("✅"):
                st.success(msg)
            elif msg.startswith("⚠️"):
                st.warning(msg)
            elif msg.startswith("💡") or msg.startswith("📧"):
                st.info(msg)
            elif msg.startswith("❌"):
                st.error(msg)
            else:
                st.write(msg)
    
    # ============================================================
    # PHASE 8: DISPLAY RESULTS
    # ============================================================
    
    st.header("📊 Results Summary")
    
    # Summary metrics
    total_commission = final_summary['Total_Commission'].sum()
    total_due = final_summary['Total_Due'].sum()
    
    col_total1, col_total2 = st.columns(2)
    with col_total1:
        st.metric("Total Commission", f"${total_commission:,.2f}")
    with col_total2:
        st.metric("Total Amount Due", f"${total_due:,.2f}", help="Sum of positive commission amounts only")
    
    st.divider()
    
    # Individual salesperson metrics
    cols = st.columns(len(final_summary))
    for idx, (_, row) in enumerate(final_summary.iterrows()):
        with cols[idx]:
            st.metric(
                row['Salesperson'],
                f"${row['Total_Commission']:,.2f}",
                delta=None
            )
    
    # Tabs for detailed views
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Commission Summary", "💰 By Category", "🏢 Revenue by Client", "📋 Full Ledger"])
    
    with tab1:
        st.subheader("Commission Summary")
        
        # Group by salesperson first
        for salesperson in final_summary['Salesperson'].unique():
            sp_commissions = all_commissions[all_commissions['Salesperson'] == salesperson].copy()
            
            # Group by Client, Category, and Rate
            summary = sp_commissions.groupby(
                ['Client', 'Category', 'Commission_Rate'], 
                as_index=False
            ).agg({
                'Invoice_Amount': 'sum',
                'Commission_Amount': 'sum'
            })
            
            # Rename columns to match spreadsheet format
            summary = summary.rename(columns={
                'Client': 'Client or Resource',
                'Commission_Rate': 'Factor',
                'Invoice_Amount': 'Revenue ($)',
                'Commission_Amount': 'Commission ($)'
            })
            
            # Sort by commission amount descending
            summary = summary.sort_values('Commission ($)', ascending=False)
            
            # Calculate total
            total_comm = summary['Commission ($)'].sum()
            
            with st.expander(f"**{salesperson}** - ${total_comm:,.2f}", expanded=True):
                st.dataframe(
                    summary.style.format({
                        'Factor': '{:.1%}',
                        'Revenue ($)': '${:,.2f}',
                        'Commission ($)': '${:,.2f}'
                    }),
                    hide_index=True,
                    use_container_width=True
                )
    
    with tab2:
        st.subheader("Commission Breakdown by Category")
        for salesperson in final_summary['Salesperson'].unique():
            sp_categories = category_summary[category_summary['Salesperson'] == salesperson]
            with st.expander(f"**{salesperson}** - ${sp_categories['Commission_Amount'].sum():,.2f}"):
                st.dataframe(
                    sp_categories[['Category', 'Commission_Amount']].style.format({'Commission_Amount': '${:,.2f}'}),
                    hide_index=True
                )
    
    with tab3:
        st.subheader(f"Revenue by Client - {year}")
        st.write(f"**Total Clients:** {len(revenue_by_client)} | **Total Revenue:** ${revenue_by_client['Total_Revenue'].sum():,.2f}")
        st.dataframe(
            revenue_by_client.style.format({'Total_Revenue': '${:,.2f}', 'Transactions': '{:.0f}'}),
            height=400
        )
    
    with tab4:
        st.subheader("Full Commission Ledger")
        ledger_sorted = all_commissions.sort_values(['Invoice_Date', 'Client'], ascending=[True, True])
        ledger_display = ledger_sorted[['Salesperson', 'Client', 'Category', 'Invoice_Date', 'Invoice_Amount', 'Commission_Rate', 'Commission_Amount', 'Source']].copy()
        ledger_display = ledger_display.rename(columns={'Client': 'Client or Resource'})
        st.dataframe(
            ledger_display.style.format({
                'Invoice_Amount': '${:,.2f}',
                'Commission_Rate': '{:.2%}',
                'Commission_Amount': '${:,.2f}',
                'Invoice_Date': lambda x: x.strftime('%Y-%m-%d') if pd.notna(x) else ''
            }),
            height=400,
            hide_index=True
        )
    
    # ============================================================
    # PHASE 9: EXPORT TO EXCEL
    # ============================================================
    
    st.divider()
    st.subheader("📥 Export Report")
    
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create Excel file in memory
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Tab 1: Overall Summary with Total Due column
            overall_df = pd.DataFrame({
                'Salesperson': final_summary['Salesperson'],
                'Total Commission': final_summary['Total_Commission'],
                'Total Due': final_summary['Total_Due']
            })
            overall_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Add Totals row to Summary sheet
            workbook = writer.book
            summary_sheet = writer.sheets['Summary']
            
            last_row = len(overall_df) + 2
            summary_sheet.cell(row=last_row, column=1, value='Totals')
            summary_sheet.cell(row=last_row, column=2, value=total_commission)
            summary_sheet.cell(row=last_row, column=3, value=total_due)
            
            # Tab 2: Category Breakdown
            category_summary.to_excel(writer, sheet_name='By_Category', index=False)
            
            # Tab 3: Commission Summary (grouped by client/category/rate)
            for salesperson in final_summary['Salesperson'].unique():
                sp_commissions = all_commissions[all_commissions['Salesperson'] == salesperson].copy()
                
                comm_summary = sp_commissions.groupby(
                    ['Client', 'Category', 'Commission_Rate'], 
                    as_index=False
                ).agg({
                    'Invoice_Amount': 'sum',
                    'Commission_Amount': 'sum'
                })
                
                comm_summary = comm_summary.rename(columns={
                    'Client': 'Client or Resource',
                    'Commission_Rate': 'Factor',
                    'Invoice_Amount': 'Revenue ($)',
                    'Commission_Amount': 'Commission ($)'
                })
                
                comm_summary = comm_summary.sort_values('Commission ($)', ascending=False)
                
                sheet_name = f"{salesperson.replace(' ', '_')}_Summary"[:31]
                comm_summary.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Tab 4: Full Ledger
            ledger_sorted = all_commissions.sort_values(['Invoice_Date', 'Client'], ascending=[True, True])
            ledger_export = ledger_sorted[['Salesperson', 'Client', 'Category', 'Invoice_Date', 'Invoice_Amount', 'Commission_Rate', 'Commission_Amount', 'Source']].copy()
            ledger_export = ledger_export.rename(columns={'Client': 'Client or Resource'})
            ledger_export.to_excel(writer, sheet_name='Full_Ledger', index=False)
            
            # Tab 5: Revenue by Client
            revenue_export = revenue_by_client.reset_index()
            revenue_export = revenue_export.rename(columns={'Client_Normalized': 'Client'})
            revenue_export.to_excel(writer, sheet_name='Revenue_by_Client', index=False)
            
            # Tab 6+: Individual salesperson tabs with category breakdown above ledger
            for salesperson in final_summary['Salesperson'].unique():
                sp_commissions = all_commissions[all_commissions['Salesperson'] == salesperson].copy()
                sp_categories = category_summary[category_summary['Salesperson'] == salesperson].copy()
                sp_total = sp_categories['Commission_Amount'].sum()
                
                # Excel sheet names limited to 31 chars
                sheet_name = salesperson.replace(' ', '_')[:31]
                
                # Create new sheet
                ws = workbook.create_sheet(title=sheet_name)
                
                # Add header
                ws.cell(row=1, column=1, value='Totals')
                ws.cell(row=2, column=1, value='Salesperson')
                ws.cell(row=2, column=2, value='Category')
                ws.cell(row=2, column=3, value='Commission_Amount')
                
                # Add category breakdown
                row_num = 3
                for _, cat_row in sp_categories.iterrows():
                    ws.cell(row=row_num, column=1, value=salesperson)
                    ws.cell(row=row_num, column=2, value=cat_row['Category'])
                    ws.cell(row=row_num, column=3, value=cat_row['Commission_Amount'])
                    row_num += 1
                
                # Add total row
                ws.cell(row=row_num, column=2, value='Total Due')
                ws.cell(row=row_num, column=3, value=max(0, sp_total))
                row_num += 2
                
                # Add ledger header
                ws.cell(row=row_num, column=1, value='Ledger')
                row_num += 1
                
                # Add ledger columns
                sp_commissions_sorted = sp_commissions.sort_values(['Invoice_Date', 'Client'], ascending=[True, True])
                sp_ledger = sp_commissions_sorted[['Client', 'Category', 'Invoice_Date', 'Invoice_Amount', 'Commission_Rate', 'Commission_Amount', 'Source']].copy()
                sp_ledger = sp_ledger.rename(columns={'Client': 'Client or Resource'})
                
                # Write ledger to sheet
                for r in dataframe_to_rows(sp_ledger, index=False, header=True):
                    for col_idx, value in enumerate(r, 1):
                        ws.cell(row=row_num, column=col_idx, value=value)
                    row_num += 1
        
        # Get the Excel data
        excel_data = output.getvalue()
        report_timestamp = datetime.now().strftime('%Y-%m-%d_%H%M')
        filename = f"Commission_Report_{year}_{report_timestamp}.xlsx"
        
        # Store for email
        st.session_state.commission_report_data['excel_file'] = excel_data
        st.session_state.commission_report_data['filename'] = filename
        
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
    except Exception as e:
        st.error(f"❌ Export failed: {e}")
        import traceback
        st.code(traceback.format_exc())

else:
    st.info("👆 Click the button above to calculate commissions for the selected year")
    
    with st.expander("ℹ️ How it works"):
        st.markdown("""
        This calculator:
        1. Loads commission rules from **Voyage_Global_Config** Google Sheet
        2. Pulls **QuickBooks** consulting income (cash basis)
        3. Pulls **BigTime** time entries for delivery and referral commissions
        4. Calculates commissions based on date ranges and rates
        5. Applies offsets (salaries, benefits, etc.)
        6. Exports detailed report to Excel with email capability
        
        **Commission Types:**
        - **Client Commission** - % of revenue from specific clients
        - **Delivery Commission** - % of own billable work
        - **Referral Commission** - % of referred staff's work
        - **Offsets** - Salaries, benefits, prior payments
        """)

# Email functionality - placed at end so it's always evaluated
if 'commission_report_data' in st.session_state:
    st.sidebar.markdown("---")
    st.sidebar.subheader("📧 Email Report")
    
    email_to = st.sidebar.text_input(
        "Send to:",
        placeholder="email@example.com",
        key="commission_email_input"
    )
    
    send_clicked = st.sidebar.button("Send Email", type="primary", use_container_width=True, key="send_commission_email")
    
    if send_clicked:
        if not email_to:
            st.sidebar.error("Enter an email address")
        else:
            try:
                from googleapiclient.discovery import build
                from google.oauth2 import service_account
                import base64
                from email.mime.multipart import MIMEMultipart
                from email.mime.base import MIMEBase
                from email.mime.text import MIMEText
                from email import encoders
                
                rd = st.session_state.commission_report_data
                
                creds = service_account.Credentials.from_service_account_info(
                    st.secrets["SERVICE_ACCOUNT_KEY"],
                    scopes=['https://www.googleapis.com/auth/gmail.send'],
                    subject='astudee@voyageadvisory.com'
                )
                
                gmail = build('gmail', 'v1', credentials=creds)
                
                msg = MIMEMultipart()
                msg['From'] = 'astudee@voyageadvisory.com'
                msg['To'] = email_to
                msg['Subject'] = f"Commission Report - {rd['year']}"
                
                # Build summary for email body
                total_comm = rd['final_summary']['Total_Commission'].sum()
                total_due = rd['final_summary']['Total_Due'].sum()
                
                body = f"""Commission Report for {rd['year']}

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

Summary:
- Total Commission: ${total_comm:,.2f}
- Total Amount Due: ${total_due:,.2f}

Breakdown by Salesperson:
"""
                for _, row in rd['final_summary'].iterrows():
                    body += f"- {row['Salesperson']}: ${row['Total_Commission']:,.2f}\n"
                
                body += "\nBest regards,\nVoyage Advisory"
                
                msg.attach(MIMEText(body, 'plain'))
                
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(rd['excel_file'])
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename={rd["filename"]}')
                msg.attach(part)
                
                raw = base64.urlsafe_b64encode(msg.as_bytes()).decode('utf-8')
                result = gmail.users().messages().send(userId='me', body={'raw': raw}).execute()
                
                st.sidebar.success(f"✅ Sent to {email_to}!")
                
            except Exception as e:
                st.sidebar.error(f"❌ {type(e).__name__}")
                st.sidebar.code(str(e))
